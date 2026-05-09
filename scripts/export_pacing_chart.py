"""One-off / CI helper: regenerate data/pacing_chart.json from the Excel workbook."""

from __future__ import annotations

import json
from datetime import time
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent


def time_to_seconds(t):
    if t is None:
        return None
    if isinstance(t, (int, float)):
        return float(t)
    if isinstance(t, time):
        return (
            t.hour * 3600
            + t.minute * 60
            + t.second
            + (t.microsecond or 0) / 1e6
        )
    raise TypeError(f"Unsupported cell type {type(t)!r}")


COLUMN_KEYS = [
    (6, "ten_x_500", {"zone": 5}),
    (7, "time_2k", {"zone": 5, "kind": "duration_total"}),
    (8, "split_2k", {"zone": 5, "kind": "split_500"}),
    (9, "five_x_5min", {"zone": 5}),
    (10, "four_x_2k", {"zone": 5}),
    (11, "three_x_2500", {"zone": 4}),
    (12, "time_6k", {"zone": 4, "kind": "duration_total"}),
    (13, "split_6k", {"zone": 4, "kind": "split_500"}),
    (14, "four_thirty_thirty", {"zone": 3}),
    (15, "ten_k", {"zone": 3}),
    (16, "two_x_6k_20", {"zone": 3}),
    (17, "hop", {"zone": 3}),
    (18, "split_offset_plus_15", {"zone": 2, "kind": "steady_offset"}),
    (19, "split_offset_plus_18", {"zone": 2, "kind": "steady_offset"}),
    (20, "split_offset_plus_21", {"zone": 1, "kind": "steady_offset"}),
    (21, "split_offset_plus_24", {"zone": 1, "kind": "steady_offset"}),
]


def main() -> None:
    xlsx = BASE_DIR / "USD Workout Pacing Chart.xlsx"
    ws = openpyxl.load_workbook(xlsx, data_only=True)["Sheet1"]

    workout_metadata = {}
    for col, key, meta in COLUMN_KEYS:
        hdr = ws.cell(8, col).value
        spm = ws.cell(7, col).value
        workout_metadata[key] = {
            "header": str(hdr) if hdr is not None else key,
            "column": col,
            "spm_hint": spm,
            **meta,
        }

    rows_out = []
    for r in range(11, 1000):
        w = ws.cell(r, 2).value
        if w is None or not isinstance(w, (int, float)):
            break
        pace_dec = ws.cell(r, 3).value
        sec500 = ws.cell(r, 4).value
        workouts = {}
        for col, key, _ in COLUMN_KEYS:
            workouts[key] = time_to_seconds(ws.cell(r, col).value)
        rows_out.append(
            {
                "row_index": r,
                "watts": float(w),
                "pace_ratio": float(pace_dec) if pace_dec is not None else None,
                "pace_seconds_500": float(sec500) if sec500 is not None else None,
                "time_2k_seconds": workouts.get("time_2k"),
                "split_2k_seconds": workouts.get("split_2k"),
                "workouts": workouts,
            }
        )

    payload = {
        "title": "USD Workout Pacing Chart",
        "source_file": "USD Workout Pacing Chart.xlsx",
        "sheet": "Sheet1",
        "zone_summary_row": {
            "zone_5": "Anaerobic (500-2000m) — 2k pace +0–5s",
            "zone_4": "Aer/Anaerobic (8'–20')",
            "zone_3": "Anaerobic threshold (30'–60')",
            "zone_2": "Aerobic (40'–90')",
            "zone_1": "Active recovery / very easy",
        },
        "workout_types": workout_metadata,
        "default_steady_workout_key": "split_offset_plus_18",
        "rows": rows_out,
    }

    out = BASE_DIR / "data" / "pacing_chart.json"
    out.parent.mkdir(exist_ok=True)
    out.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {out} ({len(rows_out)} rows)")


if __name__ == "__main__":
    main()
